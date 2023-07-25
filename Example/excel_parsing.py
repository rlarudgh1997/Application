import argparse
import csv
import os
import pickle
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
from pandas import DataFrame
# 하위 경로 import 방법 3가지
# import source.def_functon as func
# from source import def_functon as func
from source.def_functon import func_name, func_name1, isFuntionName

# from operator import length_hint


# 상위 경로 import 방법
# import sys, os
# sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
# from parent_folder import module_name


# 파일명, 시트명
filePath = "../Example/SOC_Gauge.xlsx"
sheetName = ["Description", "Privates", "Telltales", "Constants", "Events", "Sounds", "Inters", "Outputs"]
sheetInfo = list()
excelInfo = dict()

excelInfo["sheetName"] = sheetName
excelInfo["sheetInfo"] = list()


def readFromExcel(fileName, sheetName, state) :
    # Daraframe형식으로 엑셀 파일 읽기
    # df = pd.read_excel(file_name, header = None, sheet_name=["Sheet1", "Sheet2", "Sheet3"])

    # print("\n\n\n>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    # print("\t Parsing Start :", fileName, ", ", sheetName, ", ", len(sheetInfo), ", ", type(sheetInfo))

    df = pd.read_excel(fileName, sheet_name=sheetName)

    index = 0
    if (state) :
        for sheet in sheetName:
            sheetInfo.append(df[sheetName[index]])
            sheetInfo[index] = sheetInfo[index].fillna("")  # 공백 문자(NaN)를 실제 공백으로 변경
            # print("\n==================================================================================")
            print(sheetInfo[index])
            index += 1
    else :
        for sheet in excelInfo["sheetName"]:
            excelInfo["sheetInfo"].append(df[excelInfo["sheetName"][index]])
            excelInfo["sheetInfo"][index] = excelInfo["sheetInfo"][index].fillna("")  # 공백 문자(NaN)를 실제 공백으로 변경
            # print("\n==================================================================================")
            # print(excelInfo["sheetInfo"][index])
            index += 1
    # print("\n\t Parsing Complete")
    # print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<\n\n")





def editExcelData() :
    if (len(sheetInfo[0])) > 0 :
        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
        print("\t Edit Start")
        print("\n==================================================================================")
        sheetInfo[0].loc[6] =["KKH1", "", "", "KKH4", "KKH5"]
        print(sheetInfo[0])
        print("\n\t Edit Complete")
        print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<\n\n")
    else :
        print("Fail to sheetInfo len :", len(excelInfo["sheetInfo"]))




def writeToText(state, path) :
    if (state) :
        # print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
        # print("\t Write Start :", path)
        # print("\n==================================================================================")
        index = 0
        for sheet in sheetInfo:
            saveSheet = pd.DataFrame(sheet)
            # sheetFileName = "../deploy_x86/SFC/" + str(index) + "_" + sheetName[index] + ".txt"
            sheetFileName = path + str(index) + "_" + sheetName[index] + ".txt"
            # print("SaveInfo :", sheetFileName)
            writeSheet = saveSheet.to_csv(sheetFileName, sep="	", index = False)
            index += 1

        # convert_sheet1 = pd.DataFrame(sheetInfo[0])
        # print(convert_sheet1)
        # # write_sheet1 = convert_sheet1.to_csv("../Example/sheet1.txt", sep="	", index = False)
        # write_sheet1 = convert_sheet1.to_csv("../deploy_x86/SFC/sheet1.txt", sep="	", index = False)
        # print("==================================================================================")
        # print("\n\t Save Complete")
        # print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<\n\n")
        # test_list = ['one', 'two', 'three']
        # for i in sheet1.loc[0]:
        #     # print("len :", len(i), ", ", i)
        #     print(i)


    index = 0
    with pd.ExcelWriter("../Example/test_convert.xlsx") as w:
        sheetInfo[index].to_excel(w, sheet_name = sheetName[index], index = False)
        index += 1

    # 셀 병합
    # wb = Workbook()
    # ws = wb.active
    # ws.merge_cells("A1:C1")
    # wb.save("../Example/kkh.xlsx")

    wb = load_workbook("../Example/test_convert.xlsx")
    ws = wb[sheetName[0]]
    ws.merge_cells("A2:B3")
    wb.save("../Example/kkh.xlsx")



def readFromText(path, saveFilePath) :
    index = 0
    readData = list()
    for sheet in excelInfo["sheetName"]:
        filePath = path + str(index) + "_" + sheet + ".ini"
        readData.append(pd.read_csv(filePath, sep = "\t"))
        index += 1

    writeToExcel(readData, saveFilePath)



def writeToExcel(data, filePath) :
    with pd.ExcelWriter(filePath) as write:
        count = 0
        for sheet in excelInfo["sheetName"]:
            data[count].to_excel(write, sheet_name = sheetName[count], index = False)
            print("Sheet[", count, "] : ", sheetName[count])
            count += 1
        print("FilePath : ", filePath, "\n\n")


def main(argv) :
    print("\n=================================================================================================")
    print("Start Python\n")
    if len(argv) != 4 :
        print("input length error !!!!!!!!!!")
        return

    print("Argv :", len(argv), ", ", argv)
    pathDir = argv[1] + "TC/"
    pathFile = argv[1] + argv[2]
    openState = (argv[3] == "read")

    if (os.path.isdir(pathDir) == False) :
        os.mkdir(pathDir)

    print("Dir  :", os.path.isdir(pathDir), ", ", pathDir)
    print("File :", os.path.isfile(pathFile), ", ", pathFile)

    if openState :
        print("File Open !!!!!")
        if (os.path.isfile(pathFile)) :

            # # excel_parsing(filePath, sheetName, True)
            readFromExcel(pathFile, sheetName, True)
            # # editExcelData()
            writeToText(True, pathDir)

            # # func.isFuntionName()
            # # print("FuncName :", func.func_name)

            # isFuntionName()
            # print("0 FuncName :", func_name)
            # print("1 FuncName :", func_name1)

            # print("excelInfo :", excelInfo, ", len :", len(excelInfo["sheetName"]), ", ", len(excelInfo["sheetInfo"]))
            # print("sheetName :", excelInfo["sheetName"])
            # print("sheetInfo :", excelInfo["sheetInfo"])

    else :
        print("File Write !!!!!")
        readFromText(pathDir, pathFile)


    print("\n=================================================================================================")



if __name__ == '__main__' :
    argv = sys.argv
    main(argv)
