import argparse
import csv
import os
import pickle
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pandas import DataFrame
# 하위 경로 import 방법 3가지
# import source.def_functon as func
# from source import def_functon as func
# from source.def_functon import func_name, func_name1, isFuntionName

# from operator import length_hint


# 상위 경로 import 방법
# import sys, os
# sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
# from parent_folder import module_name


# 파일명, 시트명
# filePath = "../Example/SOC_Gauge.xlsx"
sheetName = list()
sheetInfo = list()
configInfo = dict()


def readConfigSetting() :
    # print(__file__)
    # print(os.path.realpath(__file__))
    # print(os.getcwd())
    currentPath = os.path.dirname(os.path.realpath(__file__))
    configFile = currentPath + "/Application.ini"

    file = open(configFile, "r")

    print("\n==================================================================================")
    datas = file.readlines()
    for data in datas :
        config = data.split("=")
        if len(config) == 2 :
            # print(config)
            key = config[0]
            value = config[1].split("\n")
            configInfo[key] = ""
            if len(value) == 2 :
                configInfo[key] = value[0]

    readSheetName = configInfo["ConfigTypeSheetName"].split(", ")
    for name in readSheetName :
        sheetName.append(name)

    file.close()




def readFromExcel(fileName, sheetName) :
    df = pd.read_excel(fileName, sheet_name=sheetName)
    index = 0
    for sheet in sheetName:
        sheetInfo.append(df[sheetName[index]])
        # sheetInfo[index] = sheetInfo[index].fillna("")  # 공백 문자(NaN)를 실제 공백으로 변경
        print(sheetInfo[index])
        index += 1





def editExcelData() :
    if (len(sheetInfo[0])) > 0 :
        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
        print("\t Edit Start")
        print("\n==================================================================================")
        sheetInfo[0].loc[6] = ["KKH1", "", "", "KKH4", "KKH5"]
        print(sheetInfo[0])
        print("\n\t Edit Complete")
        print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<\n\n")
    else :
        print("Fail to sheetInfo len :", len(sheetName))




def writeToText(path) :
    index = 0
    for sheet in sheetInfo:
        saveSheet = pd.DataFrame(sheet)
        sheetFileName = path + str(index) + "_" + sheetName[index] + ".txt"
        # writeSheet = saveSheet.to_csv(sheetFileName, sep="\t", index = False)
        saveSheet.to_csv(sheetFileName, sep="\t", index = False)
        index += 1



def readFromText(path, saveFilePath) :
    print("\n=================================================================================================")
    print("path : ", path)

    index = 0
    readData = list()
    sheetRowCount = dict()
    mergeInfoList = dict()

    for sheet in sheetName:
        filePath = path + str(index) + "_" + sheet + ".ini"
        read = pd.read_csv(filePath, sep = "\t")
        print("*************************************************************************************************")
        print("[", sheet, "] Row : ", len(read.index), ", Column : ", len(read.columns))

        startIndexList = list()
        tcNameList = list()
        vehicleTypeList = list()
        resultList = list()
        caseList = list()

        rowCount = len(read.index)
        columnCount = len(read.columns)
        sheetRowCount[sheet] = rowCount

        if sheet != "Description":
            for rowIndex in range(0, rowCount):
                for columnIndex in range(0, columnCount):
                    readStr = read.iloc[rowIndex][columnIndex]
                    if columnIndex < 4:    # 0:TCName, 1:VehicleType, 2:Result, 3:Case
                        if pd.isna(read.iloc[rowIndex][columnIndex]) == False:
                            startIndexList.append(dict({columnIndex: rowIndex}))
                    if readStr == "ExcelSplit":
                        read.iloc[rowIndex][columnIndex] = ""
                        print("Read[", rowIndex, ",", columnIndex, "] : ", readStr, " -> ", read.iloc[rowIndex][columnIndex])


            for info in startIndexList:
                keys = list(info.keys())
                for key in keys:
                    if key == 0:    # 0:TCName
                        tcNameList.append(info[key])
                    elif key == 2:    # 2:Result
                        resultList.append(info[key])
                    elif key == 3:    # 3:Case
                        caseList.append(info[key])
                    else:    # 1:VehicleType
                        vehicleTypeList.append(info[key])

            sheetMergeInfoList = dict({"TCName": tcNameList, "VehicleType": tcNameList, "Result": resultList, "Case": caseList})
            print("    TCName      : ", sheetMergeInfoList["TCName"])
            print("    VehicleType : ", sheetMergeInfoList["VehicleType"])
            print("    Result      : ", sheetMergeInfoList["Result"])
            print("    Case        : ", sheetMergeInfoList["Case"])

            mergeInfoList[sheet] = sheetMergeInfoList
            print("    SheetMergeInfo : ", sheetMergeInfoList, "\n")


        # 특정 공백 스트링(ExcelSplit) -> 실제 공백으로 제거
        read.replace("ExcelSplit", "")
        # read.str.replace("ExcelSplit", "")
        # read.replace(" ", "")


        if configInfo["ConfigTypeDeleteFileTC"] == "true" :
            print("Delete File :", filePath)
            os.remove(filePath)

        readData.append(read)
        index += 1

    writeToExcel(readData, saveFilePath)
    writeToMergeCell(sheetRowCount, mergeInfoList, saveFilePath)




def writeToExcel(data, filePath) :
    print("\n=================================================================================================")
    with pd.ExcelWriter(filePath) as write:
        count = 0
        for sheet in sheetName:
            data[count].to_excel(write, sheet_name = sheetName[count], index = False)
            # print("Sheet[", count, "] : ", sheetName[count])
            count += 1
        print("SaveFilePath : ", filePath, "\n\n")

    ## [시트 row, column 높이, 너비 자동정렬 추가]
    # wb = load_workbook(filePath)
    # ws = wb.active
    # for sheet in sheetInfo:
    #     cellAutoFit = pd.DataFrame(sheet)

def writeToMergeCell(sheetRowCount, mergeInfoData, filePath) :
    print("\n\n>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    print("[MergeInfo]")
    wb = load_workbook(filePath)

    for sheet in mergeInfoData:
        print("    [", sheet, "] : ", mergeInfoData[sheet])
        ws = wb[sheet]
        rowCount = sheetRowCount[sheet]

        for tittle in mergeInfoData[sheet]:
            titleList = mergeInfoData[sheet]
            count = len(titleList[tittle])
            if count > 0 :
                startIndexList = titleList[tittle]
                # print("        [", sheet, "-", tittle, "] : ", startIndexList)
                # print("        [", tittle, "] : ", startIndexList, ", count : ", count)
                titleName = str(tittle)
                cellMark = ""
                if titleName == "TCName":
                    cellMark = "A"
                elif titleName == "VehicleType":
                    cellMark = "B"
                elif titleName == "Result":
                    cellMark = "C"
                else:   # Case
                    cellMark = "D"


                for index in range(0, count) :
                    start = startIndexList[index]
                    if count == 1:
                        end = rowCount   # ROW 최대 값까지 병합
                    elif index == (count - 1):
                        end = rowCount
                    else:
                        end = startIndexList[index + 1]  # 리스트에 있는 다음 값까지 병합

                    gap = end - start
                    if gap > 1:
                        mergeInfo = cellMark + str(start + 2) + ":" + cellMark + str(end -1 + 2)
                        # print("            MergeCell = ", mergeInfo)
                        ws.merge_cells(mergeInfo)

    wb.save(filePath)
    # wb.save(filePath + ".MergeCell.xlsx")
    print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<\n\n")




def main(argv) :
    print("\n=================================================================================================")
    print("[Start Python]\n")
    print("Argv :", len(argv), ", ", argv)
    if len(argv) != 4 :
        print("input length error !!!!!!!!!!")
        raise

    pathDir = argv[1] + "TC/"
    pathFile = argv[1] + argv[2]
    openState = (argv[3] == "read")

    if (os.path.isdir(pathDir) == False) :
        os.mkdir(pathDir)

    print("Dir  :", os.path.isdir(pathDir), ", ", pathDir)
    print("File :", os.path.isfile(pathFile), ", ", pathFile)

    readConfigSetting()

    if openState :
        print("File - Open !!!!!")
        if (os.path.isfile(pathFile)) :
            readFromExcel(pathFile, sheetName)
            writeToText(pathDir)
        else :
            print("Fail to fiel exists : ", pathFile)
            raise

    else :
        print("File - Write !!!!!")
        readFromText(pathDir, pathFile)


    print("\n=================================================================================================")



if __name__ == '__main__' :
    argv = sys.argv
    main(argv)
