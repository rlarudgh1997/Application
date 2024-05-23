import os
import sys
import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple


class ExcelParser:
    def __init__(self, config_path="Application.ini"):
        self.sheet_names = []
        self.sheet_info = []
        self.config_info = {}
        self.config_path = config_path
        self._read_config_setting()

    def _read_config_setting(self):
        current_path = os.path.dirname(os.path.realpath(__file__))
        config_file = os.path.join(current_path, self.config_path)

        with open(config_file, "r") as file:
            for data in file.readlines():
                config = data.split("=")
                if len(config) == 2:
                    key, value = config[0].strip(), config[1].strip()
                    self.config_info[key] = value

        self.sheet_names = self.config_info["ConfigTypeSheetName"].split(", ")

    def read_from_excel(self, file_path):
        excel_merge_text_start = self.config_info["ConfigTypeExcelMergeTextStart"]
        excel_merge_text_end = self.config_info["ConfigTypeExcelMergeTextEnd"]
        excel_merge_text = self.config_info["ConfigTypeExcelMergeText"]

        wb = load_workbook(file_path, data_only=True)
        for sheet in self.sheet_names:
            current_sheet = wb[sheet]
            merged_cells = current_sheet.merged_cells.ranges
            sheet_data = []

            for row in current_sheet.iter_rows(min_row=1, max_row=current_sheet.max_row, min_col=1, max_col=current_sheet.max_column):
                data = []
                for current_cell in row:
                    read_cell_text = current_cell.value
                    check_merged = any(current_cell.coordinate in merged_cell for merged_cell in merged_cells)

                    if check_merged:
                        merged_cell = next(merged_cell for merged_cell in merged_cells if current_cell.coordinate in merged_cell)
                        row_start, column_start = merged_cell.min_row, merged_cell.min_col
                        row_end, column_end = merged_cell.max_row, merged_cell.max_col
                        current_merge_cell = current_sheet.cell(row=row_start, column=column_start)

                        if current_cell.row == row_start:
                            merge_cell_text = excel_merge_text_start
                        elif current_cell.row == row_end:
                            merge_cell_text = excel_merge_text_end
                        else:
                            merge_cell_text = excel_merge_text

                        merge_cell_text += str(current_merge_cell.value or "")
                        read_cell_text = merge_cell_text

                    data.append(read_cell_text)
                sheet_data.append(data)
            self.sheet_info.append(sheet_data)

    def write_to_text(self, path):
        for sheet_index, sheet in enumerate(self.sheet_info):
            save_sheet = pd.DataFrame(sheet)
            sheet_file_name = os.path.join(path, f"{sheet_index}_{self.sheet_names[sheet_index]}.fromExcel")
            save_sheet.to_csv(sheet_file_name, sep="\t", index=False)

    def read_from_text(self, path, save_file_path):
        excel_merge_text_start = self.config_info["ConfigTypeExcelMergeTextStart"]
        excel_merge_text_end = self.config_info["ConfigTypeExcelMergeTextEnd"]
        excel_merge_text = self.config_info["ConfigTypeExcelMergeText"]

        read_data = []
        merge_info_list = {}

        for sheet_index, sheet in enumerate(self.sheet_names):
            file_path = os.path.join(path, f"{sheet_index}_{sheet}.toExcel")
            read = pd.read_csv(file_path, sep="\t")

            row_count, column_count = len(read.index), len(read.columns)

            merge_info = []
            row_start = row_end = -1

            for column_index in range(column_count):
                for row_index in range(row_count):
                    read_text = read.iloc[row_index, column_index]

                    if isinstance(read_text, str):
                        start_text = read_text.split(excel_merge_text_start)
                        end_text = read_text.split(excel_merge_text_end)
                        merge_text = read_text.split(excel_merge_text)

                        if len(start_text) == 2:
                            row_start = row_index
                            row_end = -1
                        elif len(end_text) == 2:
                            row_end = row_index

                    if row_start >= 0 and row_end >= 0:
                        merge_info.append({column_index: sorted([row_start, row_end])})
                        row_start = row_end = -1

            merge_info_list[sheet] = merge_info

            for row_index in range(row_count):
                for column_index in range(column_count):
                    read_text = read.iloc[row_index, column_index]
                    text = ""

                    if isinstance(read_text, str):
                        start_text = read_text.split(excel_merge_text_start)
                        end_text = read_text.split(excel_merge_text_end)
                        merge_text = read_text.split(excel_merge_text)

                        if len(start_text) == 2:
                            text = start_text[1]
                        elif len(end_text) == 2:
                            text = ""
                        elif len(merge_text) == 2:
                            text = ""
                        else:
                            text = read_text
                    else:
                        text = read_text

                    read.iloc[row_index, column_index] = text

            read_data.append(read)

        self._write_to_excel(read_data, save_file_path)
        self._write_to_merge_cell(merge_info_list, save_file_path)

    def _write_to_excel(self, data, file_path):
        with pd.ExcelWriter(file_path) as writer:
            for count, sheet in enumerate(self.sheet_names):
                data[count].to_excel(writer, sheet_name=sheet, index=False)

    def _write_to_merge_cell(self, merge_info_list, file_path):
        wb = load_workbook(file_path)

        for sheet_key, sheet_value in merge_info_list.items():
            ws = wb[sheet_key]
            for merge in sheet_value:
                for column_key, row_value in merge.items():
                    if len(row_value) == 2:
                        column_string = chr(65 + column_key)
                        row_start = f"{column_string}{row_value[0] + 2}"
                        row_end = f"{column_string}{row_value[1] + 2}"
                        merge_cell = f"{row_start}:{row_end}"
                        ws.merge_cells(merge_cell)

        wb.save(file_path)


def main(argv):
    if len(argv) != 4:
        print("Usage: script.py <directory> <file_name> <read/write>")
        sys.exit(1)

    start_time = time.time()
    path_dir = os.path.join(argv[1], "TC")
    os.makedirs(path_dir, exist_ok=True)
    path_file = os.path.join(argv[1], argv[2])
    open_state = argv[3] == "read"

    parser = ExcelParser()

    if open_state:
        if os.path.isfile(path_file):
            parser.read_from_excel(path_file)
            parser.write_to_text(path_dir)
        else:
            print(f"Excel file not found: {path_file}")
            sys.exit(1)
    else:
        parser.read_from_text(path_dir, path_file)

    execution_time = time.time() - start_time
    print(f"Execution time: {execution_time:.2f} seconds")


if __name__ == "__main__":
    main(sys.argv)
