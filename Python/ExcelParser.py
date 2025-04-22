import os
import sys
import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, range_boundaries


class ExcelParser:
    def __init__(self, config_path="Application.ini"):
        self.sheet_names = []
        self.title_desc = []
        self.title_config = []
        self.title_dependent = []
        self.title_other = []
        self.sheet_info = []
        self.config_info = {}
        self.config_path = config_path
        self.read_config_setting()

    def read_config_setting(self):
        current_path = os.path.dirname(os.path.realpath(__file__))
        config_file = os.path.join(current_path, self.config_path)

        with open(config_file, "r") as file:
            for data in file.readlines():
                config = data.split("=")
                if len(config) == 2:
                    key, value = config[0].strip(), config[1].strip()
                    self.config_info[key] = value

        self.sheet_names = self.config_info["ConfigTypeSheetName"].split(", ")
        self.title_desc = self.config_info["ConfigTypeDescTitle"].split(", ")
        self.title_config = self.config_info["ConfigTypeConfigTitle"].split(", ")
        self.title_dependent = self.config_info["ConfigTypeDependentOnTitle"].split(", ")
        self.title_other = self.config_info["ConfigTypeOtherTitle"].split(", ")

    def read_from_excel(self, file_path):
        excel_merge_start = self.config_info["ConfigTypeExcelMergeStart"]
        excel_merge_end = self.config_info["ConfigTypeExcelMergeEnd"]
        excel_merge = self.config_info["ConfigTypeExcelMerge"]

        # start_time = time.time()
        wb = load_workbook(os.path.abspath(file_path), data_only=True)
        # check_time = time.time() - start_time
        # print(f"\t [CheckTime] load_workbook : {check_time:.5f} seconds")

        for sheet in self.sheet_names:
            start_time = time.time()
            sheet_data = []

#if 1    // USE_SHEET_COLUMN_OLD
            if (sheet not in wb.sheetnames):
                if (sheet == "Configs"):
                    sheet_data.append(self.title_config)
                elif (sheet == "DependentOn"):
                    sheet_data.append(self.title_dependent)
                else:
                    continue

                print("Append Sheet :", sheet, sheet_data)
                self.sheet_info.append(sheet_data)
                continue
#endif

            current_sheet = wb[sheet]

            merged_cells_dict = {}
            for merged_cell in current_sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_cells_dict[current_sheet.cell(row=row, column=col).coordinate] = merged_cell

            for row in current_sheet.iter_rows(min_row=1, max_row=current_sheet.max_row, min_col=1, max_col=current_sheet.max_column):
                data = []
                for current_cell in row:
                    # start_time_sheet_sub = time.time()

                    read_cell_text = current_cell.value

                    # check_time_any = time.time() - start_time_sheet_sub
                    # start_time_sheet_sub = time.time()

                    if current_cell.coordinate in merged_cells_dict:
                        merged_cell = merged_cells_dict[current_cell.coordinate]
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
                        row_start, column_start = min_row, min_col
                        row_end, column_end = max_row, max_col

                        current_merge_cell = current_sheet.cell(row=row_start, column=column_start)
                        if current_cell.row == row_start:
                            merge_cell_text = excel_merge_start
                        elif current_cell.row == row_end:
                            merge_cell_text = excel_merge_end
                        else:
                            merge_cell_text = excel_merge

                        merge_cell_text += str(current_merge_cell.value or "")
                        read_cell_text = merge_cell_text

                    # check_time_merge_info = time.time() - start_time_sheet_sub
                    # print("\t\t [CheckTime]", sheet, f" {check_time_any:.5f}, {check_time_merge_info:.5f} seconds")

                    data.append(read_cell_text)
                sheet_data.append(data)
            self.sheet_info.append(sheet_data)

            check_time = time.time() - start_time
            print("\t [CheckTime]", sheet, f": {check_time:.5f} seconds")

    def write_to_text(self, path):
        start_time = time.time()

        for sheet_index, sheet in enumerate(self.sheet_info):
            save_sheet = pd.DataFrame(sheet)
            sheet_file_name = os.path.join(path, f"{sheet_index}_{self.sheet_names[sheet_index]}.fromExcel")
            save_sheet.to_csv(sheet_file_name, sep="\t", index=False)

        check_time = time.time() - start_time
        print("\t [CheckTime] WriteToExcel :", f" {check_time:.5f} seconds")

    def read_from_text(self, path, save_file_path):
        excel_merge_start = self.config_info["ConfigTypeExcelMergeStart"]
        excel_merge_end = self.config_info["ConfigTypeExcelMergeEnd"]
        excel_merge = self.config_info["ConfigTypeExcelMerge"]

        # print("\n\n\t read_from_text :", save_file_path)

        read_data = []
        merge_info_list = {}

        for sheet_index, sheet in enumerate(self.sheet_names):
            file_path = os.path.join(path, f"{sheet_index}_{sheet}.toExcel")
            # print("\t read_from_text :", file_path)

            read = pd.read_csv(file_path, sep="\t")

            row_count, column_count = len(read.index), len(read.columns)

            merge_info = []
            row_start = row_end = -1

            for column_index in range(column_count):
                for row_index in range(row_count):
                    read_text = read.iloc[row_index, column_index]

                    if isinstance(read_text, str):
                        start_text = read_text.split(excel_merge_start)
                        end_text = read_text.split(excel_merge_end)
                        merge_text = read_text.split(excel_merge)

                        if len(start_text) == 2:
                            row_start = row_index
                            row_end = -1
                        elif len(end_text) == 2:
                            row_end = row_index

                    if row_start >= 0 and row_end >= 0:
                        merge_info.append({column_index: sorted([row_start, row_end])})
                        row_start = row_end = -1

            merge_info_list[sheet] = merge_info

            for row_index in range(row_count):
                for column_index in range(column_count):
                    read_text = read.iloc[row_index, column_index]
                    text = ""

                    if isinstance(read_text, str):
                        start_text = read_text.split(excel_merge_start)
                        end_text = read_text.split(excel_merge_end)
                        merge_text = read_text.split(excel_merge)

                        if len(start_text) == 2:
                            text = start_text[1]
                        elif len(end_text) == 2:
                            text = ""
                        elif len(merge_text) == 2:
                            text = ""
                        else:
                            text = read_text
                    else:
                        text = read_text

                    read.iloc[row_index, column_index] = text

            read_data.append(read)

        self.write_to_excel(read_data, save_file_path)
        self.write_to_merge_cell(merge_info_list, save_file_path)

    def write_to_excel(self, data, file_path):
        with pd.ExcelWriter(file_path) as writer:
            for count, sheet in enumerate(self.sheet_names):
                data[count].to_excel(writer, sheet_name=sheet, index=False)

    def write_to_merge_cell(self, merge_info_list, file_path):
        wb = load_workbook(file_path)

        for sheet_key, sheet_value in merge_info_list.items():
            ws = wb[sheet_key]
            for merge in sheet_value:
                for column_key, row_value in merge.items():
                    if len(row_value) == 2:
                        column_string = chr(65 + column_key)
                        row_start = f"{column_string}{row_value[0] + 2}"
                        row_end = f"{column_string}{row_value[1] + 2}"
                        merge_cell = f"{row_start}:{row_end}"
                        ws.merge_cells(merge_cell)

        wb.save(file_path)

    @staticmethod    # self 변수를 포함 하지 않는 함수에 선언
    def find_file_name(directory, fileName):
        for file in os.listdir(directory):
            if file.lower() == fileName.lower():
                return os.path.join(directory, file)
        return ""

def main(argv):
    # print("[ExcelParser.py]")

    if len(argv) != 4:
        print("\t Usage     : [python3 Excelparser.py <directory> <file_name> <read/write>]\n\n")
        for index in range(0, len(argv)):
            print("\t Argv[", index, "] :", argv[index])
        sys.exit(1)

    start_time = time.time()
    path_dir = os.path.join(argv[1], "TC")
    os.makedirs(path_dir, exist_ok=True)
    open_state = (argv[3] == "read")
    parser = ExcelParser()

    if open_state:
        path_file = parser.find_file_name(argv[1], argv[2])    # Read 인 경우만 대소문 구분 없이 파일명 인식 하도록
        if os.path.isfile(path_file):
            parser.read_from_excel(path_file)
            parser.write_to_text(path_dir)
        else:
            print(f"Excel file not found: {path_file}")
            sys.exit(1)
    else:
        path_file = os.path.join(argv[1], argv[2])
        parser.read_from_text(path_dir, path_file)

    execution_time = time.time() - start_time
    print(f"[CheckTime] Total : {execution_time:.5f} seconds")


if __name__ == "__main__":
    main(sys.argv)
