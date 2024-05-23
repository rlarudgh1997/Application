import argparse
import csv
import os
import pickle
import sys
import time

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import coordinate_to_tuple, range_to_tuple
# from openpyxl.utils import get_column_letter
from pandas import DataFrame

# # 하위 경로 import 방법 3가지
# import source.def_functon as func
# from source import def_functon as func
# from source.def_functon import func_name, func_name1, isFuntionName
# from operator import length_hint
# # 상위 경로 import 방법
# import sys, os
# sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
# from parent_folder import module_name


sheetName = list()
sheetInfo = list()
configInfo = dict()

def readConfigSetting() :
    # print(__file__)
    # print(os.path.realpath(__file__))
    # print(os.getcwd())
    currentPath = os.path.dirname(os.path.realpath(__file__))
    configFile = currentPath + "/Application.ini"

    file = open(configFile, "r")

    # print("\n==============================================================================================")
    datas = file.readlines()
    for data in datas :
        config = data.split("=")
        if len(config) == 2 :
            # print(config)
            key = config[0]
            value = config[1].split("\n")
            configInfo[key] = ""
            if len(value) == 2 :
                configInfo[key] = value[0]
    file.close()

    readSheetName = configInfo["ConfigTypeSheetName"].split(", ")
    for name in readSheetName :
        sheetName.append(name)

    # print("SheetName :", sheetName,      ",", type(sheetName))
    # print("==============================================================================================\n")





def readFromExcel(file_path, sheet_name):
    excelMergeTextStart = configInfo["ConfigTypeExcelMergeTextStart"]
    excelMergeTextEnd = configInfo["ConfigTypeExcelMergeTextEnd"]
    excelMergeText = configInfo["ConfigTypeExcelMergeText"]


    wb = load_workbook(file_path, data_only=True)
    for sheet in sheetName:
        currentSheet = wb[sheet]
        merged_cells = currentSheet.merged_cells.ranges
        sheetData = list()

        if False:    # new code : True, False -> 파싱 속도 오래 걸림
            for row in currentSheet.iter_rows(min_row=1, max_row=currentSheet.max_row, min_col=1, max_col=currentSheet.max_column):
                data = list()
                for currentCell in row:    # 현재 셀이 머지된 셀인지 확인
                    readCellText = currentCell.value
                    checkMerged = False

                    for merged_cell in merged_cells:
                        if currentCell.coordinate in merged_cell:
                            checkMerged = True
                            break

                    if checkMerged:    # 머지된 셀의 경우
                        rowStart, columnStart, rowEnd, columnEnd = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
                        currentMergeCell = currentSheet.cell(row=rowStart, column=columnStart)

                        if True:    # new code : True, False
                            if currentCell.row == rowStart:
                                mergeCellText = excelMergeTextStart
                            elif currentCell.row == rowEnd:
                                mergeCellText = excelMergeTextEnd
                            else:
                                mergeCellText = excelMergeText

                            if readCellText:
                                mergeCellText = mergeCellText + currentMergeCell.value
                        else:
                            if readCellText is None:
                                mergeCellText = excelMergeText
                            else:
                                mergeCellText = excelMergeText + currentMergeCell.value

                        readCellText = mergeCellText
                        # print("1 Merge[", currentCell.row, ",", currentCell.column, "] :", readCellText)

                    data.append(readCellText)
                    # print("1 Text[", currentCell.row, ",", currentCell.column, "] :", checkMerged, ",", readCellText)

                sheetData.append(data)

            sheetInfo.append(sheetData)

        else:    # 파싱 속도 빠름
            for row in currentSheet.iter_rows(min_row=1, max_row=currentSheet.max_row, min_col=1, max_col=currentSheet.max_column):
                data = list()
                for currentCell in row:
                    if (type(currentCell.value) == str) and ("\n" in currentCell.value):
                        readCellText = currentCell.value.replace("\n", "")
                        print("\t Cell with line break at -", sheet, "[", rowIndex, ",", columnIndex, "] :", readCellText)
                    else:
                        readCellText = currentCell.value

                    rowIndex, columnIndex = coordinate_to_tuple(currentCell.coordinate)

                    checkMerged = False
                    rowStart = (-1)
                    rowEnd = (-1)
                    for merged_cell in merged_cells:
                        columnStart, rowStart, columnEnd, rowEnd = merged_cell.bounds
                        if columnIndex == columnStart:
                            if rowIndex >= rowStart and rowIndex <= rowEnd:
                                checkMerged = True
                                break

                    if checkMerged:
                        if rowIndex == rowStart:
                            mergeCellText = excelMergeTextStart
                        elif rowIndex == rowEnd:
                            mergeCellText = excelMergeTextEnd
                        else:
                            mergeCellText = excelMergeText

                        if readCellText:
                            mergeCellText = mergeCellText + str(readCellText)

                        readCellText = mergeCellText
                        # print("2 Merge[", rowIndex, ",", columnIndex, "] :", readCellText, rowStart, rowEnd)

                    data.append(readCellText)
                    # print("2 Text[", rowIndex, ",", columnIndex, "] :", checkMerged, ",", readCellText)

                sheetData.append(data)

            sheetInfo.append(sheetData)





def writeToText(path) :
    sheetIndex = 0
    for sheet in sheetInfo:
        saveSheet = pd.DataFrame(sheet)
        sheetFileName = path + str(sheetIndex) + "_" + sheetName[sheetIndex] + ".fromExcel"
        # writeSheet = saveSheet.to_csv(sheetFileName, sep="\t", sheetIndex = False)
        saveSheet.to_csv(sheetFileName, sep="\t", index = False)
        sheetIndex += 1



def readFromText(path, saveFilePath) :
    # print("\n=================================================================================================")
    # print("readFromText : ", path)

    excelMergeTextStart = configInfo["ConfigTypeExcelMergeTextStart"]
    excelMergeTextEnd = configInfo["ConfigTypeExcelMergeTextEnd"]
    excelMergeText = configInfo["ConfigTypeExcelMergeText"]

    readData = list()
    mergeInfoList = dict()
    sheetIndex = (-1)

    for sheet in sheetName:
        sheetIndex += 1
        filePath = path + str(sheetIndex) + "_" + sheet + ".toExcel"
        read = pd.read_csv(filePath, sep = "\t")

        # print(sheet, ":", len(read.index), len(read.columns))

        rowCount = len(read.index)
        columnCount = len(read.columns)


        if True:    # Merge Cell 정보 구성
            rowStart = (-1)
            rowEnd = (-1)
            mergeInfo = list()

            for columnIndex in range(0, columnCount):
                for rowIndex in range(0, rowCount):
                    readText = read.iloc[rowIndex, columnIndex]

                    if type(readText) == str:
                        startText = readText.split(excelMergeTextStart)
                        endText = readText.split(excelMergeTextEnd)
                        mergeText = readText.split(excelMergeText)
                        if len(startText) == 2 :
                            rowStart = rowIndex
                            rowEnd = (-1)
                        elif len(endText) == 2 :
                            rowEnd = rowIndex

                    if (rowStart >= 0) and (rowEnd >= 0):
                        # print("\t\t Info :", columnIndex, rowStart, rowEnd)
                        info = list({rowStart, rowEnd})
                        info.sort(reverse=False)    # 오름 차순 정렬
                        mergeInfo.append(dict({columnIndex: info}))
                        rowStart = (-1)
                        rowEnd = (-1)

            # print("\t MergeInfo :", mergeInfo)
            mergeInfoList[sheet] = mergeInfo


        if True:    # "ExcelMergeText" -> "" 으로 변경
            for rowIndex in range(0, rowCount):
                for columnIndex in range(0, columnCount):
                    readText = read.iloc[rowIndex, columnIndex]
                    # readText = read.loc[rowIndex][columnIndex]
                    # readText = read.iat[rowIndex, columnIndex]
                    # readText = read.at[rowIndex][columnIndex]
                    text = ""
                    if type(readText) == str:
                        startText = readText.split(excelMergeTextStart)
                        endText = readText.split(excelMergeTextEnd)
                        mergeText = readText.split(excelMergeText)
                        typeText = ""
                        if len(startText) == 2 :
                            typeText = "Start "
                            text = startText[1]
                        elif len(endText) == 2 :
                            typeText = "End   "
                            text = ""   # endText[1]
                        elif len(mergeText) == 2 :
                            typeText = "Merge "
                            text = ""   # mergeText[1]
                        else:
                            typeText = "Normal"
                            text = readText
                    # elif type(readText) == int:
                    else:
                        typeText = "Number"
                        # text = ""
                        text = readText    # 시트내에 숫자만 있는 경우 파일로 저장 안되는 이슈 수정 : 자료형 <class 'numpy.float64'>

                    # if sheetIndex == 0 :
                    #    print("Data Type :", type(readText), "->", readText)

                    read.iloc[rowIndex, columnIndex] = text
                    # read.loc[rowIndex][columnIndex] = text
                    # read.iat[rowIndex, columnIndex] = text    # Python 3.10 : Warning Message
                    # read.at[rowIndex][columnIndex] = text
                    # print(typeText, "[", rowIndex, ",", columnIndex, "] :", text)
                    # print(sheet, "[", rowIndex, "][", columnIndex, "] :", readText, "->", text, "->", read.iloc[rowIndex, columnIndex])


            readData.append(read)



        # # Folder(TC), File(*.toExcel) -->> 삭제 동작 CPP 코드로 변경
        # if True:    # 파일 삭제
        #     if configInfo["ConfigTypeDeleteFileTC"] == "true" :
        #         os.remove(filePath)    # Delete : File


    if True:    # 파일 저장
        writeToExcel(readData, saveFilePath)


    if True:    # 저장한 파일에서 머지 정보 업데이트 후 다시 저장
        writeToMergeCell(mergeInfoList, saveFilePath)



def writeToExcel(data, filePath) :
    # print("\n=================================================================================================")
    with pd.ExcelWriter(filePath) as write:
        count = 0
        for sheet in sheetName:
            data[count].to_excel(write, sheet_name = sheetName[count], index = False)
            # print("Sheet[", count, "] : ", sheetName[count])
            count += 1
        # print("writeToExcel : ", filePath)

    ## [시트 row, column 높이, 너비 자동정렬 추가]
    # wb = load_workbook(filePath)
    # ws = wb.active
    # for sheet in sheetInfo:
    #     cellAutoFit = pd.DataFrame(sheet)




def writeToMergeCell(mergeInfoList, filePath) :
    # print("\n=================================================================================================")

    # wb = load_workbook(filePath, data_only=True)
    wb = load_workbook(filePath)

    for sheetKey, sheetValue in mergeInfoList.items():
        # print("[", sheetKey, "] :", sheetValue)

        ws = wb[sheetKey]
        for merge in sheetValue:
            for columnKey, rowValue in merge.items():
                if len(rowValue) == 2:
                    columString = chr(65 + columnKey)   # 0 ~   -->>   A ~   값으로 변경
                    rowStart = columString + str(rowValue[0] + 2)   # 엑셀 병합시 : 시작값 1부터, Row 1 은 Title 임
                    rowEnd   = columString + str(rowValue[1] + 2)   # 엑셀 병합시 : 시작값 1부터, Row 1 은 Title 임
                    mergeCell = rowStart + ":" + rowEnd
                    # print("\t Merge[", columnKey, "] :", mergeCell)
                    ws.merge_cells(mergeCell)

    wb.save(filePath)
    # print("")
    # print("writeToMergeCell : ", filePath)














def main(argv) :
    print("\n=================================================================================================")
    print("[Start Python]")
    print("  Argv :", argv)
    if len(argv) != 4 :
        print("\n\t Fail to input argument length :", len(argv))
        raise


    start_time = time.time()
    pathDir = argv[1] + "TC/"
    if (os.path.isdir(pathDir) == False) :
        os.mkdir(pathDir)
    pathFile = argv[1] + argv[2]
    openState = (argv[3] == "read")

    readConfigSetting()

    print("\n==============================================================")
    print("  Dir  :", os.path.isdir(pathDir), pathDir)
    print("  File :", os.path.isfile(pathFile), pathFile)
    print("\n==============================================================")

    if openState :
        print("[엑셀 파일 열기]\n")
        if (os.path.isfile(pathFile)) :
            readFromExcel(pathFile, sheetName)
            writeToText(pathDir)
        else :
            print("엑셀 파일이 없습니다. Path :", pathFile)
            raise

        execution_time = time.time() - start_time
        print("\n==============================================================")
        print("\t 엑셀 파일 열기 :", execution_time, "초")


    else :
        print("[엑셀 파일 쓰기]\n")
        readFromText(pathDir, pathFile)

        execution_time = time.time() - start_time
        print("\n==============================================================")
        print("\t 엑셀 파일 쓰기 :", execution_time, "초")


    print("\n=================================================================================================")





if __name__ == '__main__' :
    argv = sys.argv
    main(argv)
